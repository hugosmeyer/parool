#!/usr/bin/python3

from copy import copy
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
import os
import sys

# Check if a file exists
def check_file(filename):
    try:
        filehndl = open(filename,"r")
    except IOError:
        return False, "Unable to open type file: " + filename
    else:
        filehndl.close()
        return True, ""
    
def defnfileprse(file_path):
    ini_data = dict()

    with open(file_path, 'r') as file:
        currsect = None
        for line in file:
            line = line.strip()
            if line.startswith('[') and line.endswith(']'):
                currsect = line[1:-1]
                ini_data[currsect] = list()
            elif '=' in line:
                key, value = map(str.strip, line.split('=', 1))
                if not key == "":
                    if key is None:
                       key = ""
                    ini_data[currsect].append([key, value])
    return ini_data

def copycellfrmt(fromcell, destcell):
    destcell.font          = copy(fromcell.font)
    destcell.border        = copy(fromcell.border)
    destcell.fill          = copy(fromcell.fill)
    destcell.number_format = copy(fromcell.number_format)
    destcell.protection    = copy(fromcell.protection)
    destcell.alignment     = copy(fromcell.alignment)

def makefontbold(cell):
    thisfont = copy(cell.font)
    thisfont.bold = True
    cell.font = thisfont

def maketextwrap(cell):
    cell.alignment = Alignment(horizontal='general',vertical='center',text_rotation=0,wrap_text=True,shrink_to_fit=False,indent=0)

def makebgndgrey(cell):
    cell.fill = PatternFill(fill_type="lightGray",start_color='FFCCCCCC',end_color='FFCCCCCC')

def maketextcntr(cell):
    cell.alignment = Alignment(horizontal='general',vertical='center',text_rotation=0,wrap_text=True,shrink_to_fit=False,indent=0)

# Main, start of the program
if __name__ == "__main__":

    #Assume all files are present
    filepass = True

    busnunitname = str(sys.argv[1])
    exclfilename = str(sys.argv[2])
    cldrmnth     = str(sys.argv[3])
    cldryear     = str(sys.argv[4])
    defnfilename = busnunitname + ".ini"

    # Check if definition file exists
    filechckbool, filechckrslt = check_file(defnfilename)
    if not filechckbool:
        filepass = False
    
    # Check if Excel file exists
    filechckbool, filechckrslt = check_file(exclfilename)
    if not filechckbool:
        filepass = False
    
if not filepass:
    sys.exit(1)

# Parse the definition file
defndict = defnfileprse(defnfilename)

exclmainbook = load_workbook(exclfilename,data_only=True)

exclmainname = exclmainbook.sheetnames[0]
exclmainshet = exclmainbook[exclmainbook.sheetnames[0]]
exclrowsused = exclmainshet.max_row
exclcolsused = exclmainshet.max_column

colmcntr = 1

# Make a list of column header from the input sheet
maincolmhdrs = {}
while colmcntr <= exclcolsused:
    maincolmhdrs[exclmainshet.cell(row = 1, column = colmcntr).value] = colmcntr
    colmcntr += 1

# Assume there will be something in row1, col 1 and used that as a font reference
cellfontrrefr = exclmainshet.cell(row = 1, column = 1)

# Process each section in the INI file
for defn in defndict:
    thisdefn = defndict[defn]
    
    # Default this in case there is no SKIP
    rowsstrt = 1

    repldefn = list()
    for testname, testvalu in thisdefn:
        testskip = False
        if testname == "SKIP":
            testvalu.strip()
            if testvalu.isdigit():
                rowsstrt = int(testvalu) + 1
            testskip = True
    
        if not testskip:
            repldefn.append([testname, testvalu])

    thisdefn = repldefn
    del repldefn
    
    defnname, defntype = defn.split(".")
    
    if defntype == "FILE":
        destbook = Workbook()
        destshet = destbook.active
        destshet.title = defnname
    else:
        destshet = exclmainbook.create_sheet(title = defnname)
        titlcell = destshet.cell(row = rowsstrt, column = 1)
        titlcell.font = Font(name='Arial',size=14,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
        makefontbold(titlcell)
        titlcell.value = busnunitname + " " + defnname + " - " + cldrmnth + " " + cldryear
        rowsstrt += 2
    
    chdralgn = alignment=Alignment(horizontal='general',vertical='center',text_rotation=0,wrap_text=True,shrink_to_fit=False,indent=0)
    destcolm = 1
    nzrocols = []
    totlcols = []
    totlflag = False

    colmcntr = 1

    for defncolm in thisdefn:
        
        maincolm,thiscolm = defncolm
        
        # _NZ_ will have rows with an empty or zero value in this column to to be excluded
        if "_NZ_" in thiscolm:
            nzrocols.append(colmcntr)
            thiscolm = thiscolm.replace("_NZ_", "")
        
        # _SUM_ will have the values in this column added up and shown below
        if "_SUM_" in thiscolm:
            totlcols.append(colmcntr)
            totlflag = True
            thiscolm = thiscolm.replace("_SUM_", "")
            
        colmcntr += 1
        
        # insert the header
        destcell = destshet.cell(row = rowsstrt, column = destcolm)
        # this will alledgedly autossize the row in Excel
        destshet.row_dimensions[1].height = None

        if maincolm in maincolmhdrs:
            fromcell=exclmainshet.cell(row = 1, column = maincolmhdrs[maincolm])
            destcell.value = thiscolm
        else: 
            destcell.value = maincolm
        copycellfrmt(cellfontrrefr,destcell)
        makebgndgrey(destcell)
        maketextwrap(destcell)
        maketextcntr(destcell)

        # load the values below the header
        rowscntr = 2
        while rowscntr <= exclrowsused:
            destcell = destshet.cell(row = rowscntr + rowsstrt - 1, column = destcolm)
            destshet.row_dimensions[rowscntr].height = None
            
            if maincolm in maincolmhdrs:
                fromcell=exclmainshet.cell(row = rowscntr, column = maincolmhdrs[maincolm])
                if fromcell.has_style:
                    copycellfrmt(fromcell, destcell)
                destcell.value = fromcell.value 
            else:
                # Copy style from column one for hard coded values
                fromcell=exclmainshet.cell(row = rowsstrt+1, column = 1)
                if fromcell.has_style:
                    copycellfrmt(fromcell, destcell)
                destcell.value = thiscolm
            rowscntr += 1
        destcolm += 1

    # _NZ_ CHECK
    # Check which rows have a zero or empty value in a column marked with _NZ_
    rowscntr = rowsstrt + 1
    dlterows = []
    while rowscntr <= exclrowsused + rowsstrt:
        for colmnmbr in nzrocols:
            destcell = destshet.cell(row = rowscntr, column = colmnmbr)
            if type(destcell.value) in [int, float]:
                if destcell.value == 0:
                    if rowscntr not in dlterows:
                        dlterows.append(rowscntr)
            elif destcell.value in [None, ""]:
                if rowscntr not in dlterows:
                    dlterows.append(rowscntr)
        rowscntr += 1

    # Now remove them from the sheet
    dlterows.sort(reverse = True)
    if dlterows:
        for dlterown in dlterows:
            destshet.delete_rows(dlterown)

    # Add up the totals if there are any
    if totlflag:
        rowsused = destshet.max_row
        colsused = destshet.max_column
        print("colsused = ",colsused)
        colmnmbr = 1
        while colmnmbr <= colsused:
            totlcell = destshet.cell(row = rowsused + 1, column = colmnmbr)
            abovcell = destshet.cell(row = rowsused    , column = colmnmbr)
            copycellfrmt(abovcell, totlcell)
            makefontbold(totlcell)
            makebgndgrey(totlcell)
            
            if colmnmbr in totlcols:
                rowscntr = rowsstrt + 1
                totlvalu = 0
                while rowscntr <= rowsused:
                    valucell = destshet.cell(row = rowscntr, column = colmnmbr)
                    if type(valucell.value) in [int, float]:
                        totlvalu += valucell.value
                    rowscntr += 1

                totlcell.value = totlvalu

            if colmnmbr == 1:
                totlcell.value = "Total:"

            colmnmbr += 1

    # Make all the columns auto sizing
    for cols in destshet.columns:
        col = get_column_letter(cols[0].column)
        destshet.column_dimensions[col].auto_size = True
        
    # Save it if it is a file type
    if defntype == "FILE":
        destbook.save(filename=defnname+".xlsx")

# Save the main sheet anyway
exclmainbook.save(filename="_"+exclfilename)
sys.exit()