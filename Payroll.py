#!/usr/bin/env python3
import sys
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import datetime
from copy import copy
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
import os
from os import path
import sys
import configparser
from processFiles import processFiles

class Payroll:

    def __init__(self, root):
        self.root = root
        self.root.title("Excel File Processor")
        self.root.geometry("800x600")
        
        self.root.resizable(False, False)

        if hasattr(sys, '_MEIPASS'):
            img_path = os.path.join(sys._MEIPASS, "background.png")
        else:
            img_path = os.path.abspath("background.png")

        bg_image = Image.open(img_path)
        bg_image = bg_image.resize((800, 600) )
        self.bg_photo = ImageTk.PhotoImage(bg_image)

        #
        # Create a canvas to display the background image
        #
        canvas = tk.Canvas(root, width=800, height=600)
        canvas.pack()
        canvas.create_image(0, 0, anchor=tk.NW, image=self.bg_photo)

        self.dropdownwdth = 100
        self.filenamewdth = 400
        self.lablwdth     = 100
        self.butnwdth     = 55
        self.ypos     = 440
        self.yposincr = 30

        self.lablxpos = 10
        self.valuxpos = self.lablxpos + self.lablwdth + 10
        self.butnxpos = self.valuxpos + self.filenamewdth + 10
        # 
        # Recipe
        #
        self.rcpeflnmvalu = tk.StringVar()
        rcpeflnmlabl = tk.Label(root, text="Recipe File:")
        canvas.create_window(self.lablxpos, self.ypos, window=rcpeflnmlabl, width=self.lablwdth, anchor="nw" )

        self.rcpeflnmentr = tk.Entry(root, textvariable=self.rcpeflnmvalu)
        canvas.create_window(self.valuxpos, self.ypos, window=self.rcpeflnmentr, width=self.filenamewdth, anchor="nw")

        rcpeflnmbutn = tk.Button(root, text="Select", command=self.rcpeflnmslct)
        canvas.create_window(self.butnxpos, self.ypos, window=rcpeflnmbutn , anchor="nw", width=self.butnwdth)

        self.ypos += self.yposincr
        
        #
        # Company
        #
        compnamelabl = tk.Label(root, text="Company:", anchor="w")
        canvas.create_window(self.lablxpos, self.ypos, window=compnamelabl,width = self.lablwdth, anchor="nw")

        self.compnamevalu = tk.StringVar()
        self.compdropdown = ttk.Combobox(root, textvariable=self.compnamevalu, values=[])
        self.compdropdown.state = "disabled"
        canvas.create_window(self.valuxpos, self.ypos, window=self.compdropdown, width=self.dropdownwdth, anchor="nw")
     
        self.ypos += self.yposincr

        #
        # Month selection
        #
        mnthlabl = tk.Label(root, text="Month:", anchor="w")
        canvas.create_window(self.lablxpos, self.ypos, window=mnthlabl,width = self.lablwdth, anchor="nw")

        self.mnthnamevalu = tk.StringVar()
        self.mnthdropdown = ttk.Combobox(root, textvariable=self.mnthnamevalu, values=[
            "Jan", "Feb", "Mar", "Apr", "May", "Jun",
            "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"
        ])
        canvas.create_window(self.valuxpos, self.ypos, window=self.mnthdropdown, width=self.dropdownwdth, anchor="nw")

        self.ypos += self.yposincr

        #
        # Year selection
        #
        year_label = tk.Label(root, text="Year:", anchor="w")
        canvas.create_window(self.lablxpos, self.ypos, window=year_label,width = self.lablwdth, anchor="nw")

        self.yearnamevalu = tk.StringVar()
        self.yeardropdown = ttk.Combobox(root, textvariable=self.yearnamevalu, values=self.yearvalulist())
        canvas.create_window(self.valuxpos, self.ypos, window=self.yeardropdown, width=self.dropdownwdth, anchor="nw")

        self.ypos += self.yposincr

        #
        # Payroll File
        #
        self.prolflnmvalu = tk.StringVar()
        self.prollabl     = tk.Label(root, text="Payroll File:", anchor="w")
        canvas.create_window(self.lablxpos, self.ypos, window=self.prollabl, width=self.lablwdth , anchor="nw")

        prolflnmentr = tk.Entry(root, textvariable=self.prolflnmvalu)
        canvas.create_window(self.valuxpos, self.ypos, window=prolflnmentr, width=self.filenamewdth, anchor="nw")

        prolflnmbutn = tk.Button(root, text="Select", command=self.prolflnmslct)
        canvas.create_window(self.butnxpos, self.ypos, window=prolflnmbutn, anchor="nw", width=self.butnwdth)
                
        self.butnxpos += self.butnwdth + 10
        #
        # Process button
        #
        process_button = tk.Button(root, text="Process", command=self.process_data)
        canvas.create_window(self.butnxpos, self.ypos, window=process_button, anchor="nw", width=self.butnwdth)
        
        self.butnxpos += self.butnwdth + 10
        #
        # Exit Button
        #
        exit_button = tk.Button(root, text="Exit",    command=root.quit)
        canvas.create_window(self.butnxpos, self.ypos, window=exit_button, anchor="nw", width=self.butnwdth)

    
    def rcpeflnmslct(self):
            flnmslct = filedialog.askopenfilename(title="Select Recipe File", filetypes=[("INI files", "*.ini"), ("All files", "*.*")])
            if flnmslct:
                self.rcpeflnmvalu.set(flnmslct)
            
                print("rcpeflnmvalu = ", self.rcpeflnmvalu)
                # Parse INI file

                config = configparser.ConfigParser()
                config.read(flnmslct)

                self.complist = []
                if "COMPANIES" in config:
                    for compkeyn in config["COMPANIES"].keys():
                        self.complist.append (compkeyn + " - " + config["COMPANIES"].get(compkeyn) )
                    self.compdropdown.config(values = self.complist)
                    
    def prolflnmslct(self):
        file_path = filedialog.askopenfilename(title="Select Payroll Excel File", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")] )
        self.prolflnmvalu.set(file_path)


    def yearvalulist(self):
        current_year = datetime.datetime.now().year
        return[str(current_year - 1), str(current_year), str(current_year + 1)]
    
    
    def process_data(self):
        defnfilename = self.rcpeflnmvalu.get()
        compname     = self.compnamevalu.get()
        cldrmnth     = self.mnthnamevalu.get()
        cldryear     = self.yearnamevalu.get()
        exclfilename = self.prolflnmvalu.get()
    
        result = processFiles(defnfilename,exclfilename,cldrmnth,cldryear,compname,False)

        messagebox.showinfo("Done",result)

if __name__ == "__main__":
    root = tk.Tk()
    app = Payroll(root)
    root.mainloop()
