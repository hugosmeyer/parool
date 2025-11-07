#!/usr/bin/env python3
import os
from os import path
from copy import copy
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import logging

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)
def check_file(filename):
    """Check if a file exists and is readable.
    
    Args:
        filename: Path to the file to check
        
    Returns:
        tuple: (bool, str) - (True, "") if file exists, (False, error_message) otherwise
    """
    try:
        with open(filename, "r") as filehndl:
            pass
    except IOError:
        return False, f"Unable to open file: {filename}"
    else:
        return True, ""


def defnfileprse(file_path):
    """Parse an INI-style definition file.
    
    Parses a file with [SECTIONS] and key=value pairs into a dictionary structure.
    
    Args:
        file_path: Path to the INI file
        
    Returns:
        dict: Dictionary with section names as keys and lists of [key, value] pairs as values
    """
    ini_data = dict()

    with open(file_path, 'r') as file:
        currsect = None
        for line in file:
            line = line.strip()
            if line.startswith('[') and line.endswith(']'):
                currsect = line[1:-1]
                ini_data[currsect] = list()
            elif '=' in line:
                key, value = map(str.strip, line.split('=', 1))
                if key:  # Simplified - empty string is falsy
                    ini_data[currsect].append([key, value])
    return ini_data


# Define the underline formats for Total headers and values
BORDER_THIN = Side(border_style="thin", color="000000")
BORDER_THICK = Side(border_style="thick", color="000000")
CELL_BORDER_THIN = Border(bottom=BORDER_THIN)
CELL_BORDER_THICK = Border(bottom=BORDER_THICK)
NUMBER_FORMAT = "### ### ### ##0.00"


def makefontbold(cell):
    """Make a cell's font bold.
    
    Args:
        cell: The Excel cell to modify
    """
    thisfont = copy(cell.font)
    thisfont.bold = True
    cell.font = thisfont


def fontsizenrml(cell):
    """Set cell font to normal size (10pt Arial).
    
    Args:
        cell: The Excel cell to modify
    """
    cell.font = Font(
        name='Arial', size=10, bold=None, italic=False,
        vertAlign=None, underline=None, strike=False, color='FF000000'
    )


def fontsizelrge(cell):
    """Set cell font to large size (14pt Arial).
    
    Args:
        cell: The Excel cell to modify
    """
    cell.font = Font(
        name='Arial', size=14, bold=None, italic=False,
        vertAlign=None, underline=None, strike=False, color='FF000000'
    )


def fillcellcolr(cell):
    """Fill cell with light blue background color.
    
    Args:
        cell: The Excel cell to modify
    """
    cell.fill = PatternFill(
        fill_type="lightGray", start_color='DAE3F3', end_color='DAE3F3'
    )


def frmttotltitl(cell):
    """Format a cell as a total title (bold, colored, bottom-aligned).
    
    Args:
        cell: The Excel cell to modify
    """
    fontsizenrml(cell)
    makefontbold(cell)
    fillcellcolr(cell)
    cell.alignment = Alignment(
        horizontal='general', vertical='bottom', text_rotation=0,
        wrap_text=True, shrink_to_fit=False, indent=0
    )


def frmttotlvalu(cell):
    """Format a cell as a total value (bold, thick border, number format).
    
    Args:
        cell: The Excel cell to modify
    """
    fontsizenrml(cell)
    makefontbold(cell)
    cell.border = CELL_BORDER_THICK
    cell.number_format = NUMBER_FORMAT
    cellabov = cell.parent.cell(row=cell.row - 1, column=cell.column)
    cellabov.border = CELL_BORDER_THIN 


def populateTheSheet(exclmainshet, maincolmhdrs, destshet, thisdefn, defnname, 
                     busnunitname, cldrmnth, cldryear, totlcols, nzrocols, 
                     anzrcols, aftrtotldefn):
    """Populate a worksheet with formatted data from the main Excel sheet.
    
    Args:
        exclmainshet: Source Excel worksheet
        maincolmhdrs: Dictionary mapping column names to column indices
        destshet: Destination worksheet to populate
        thisdefn: List of [source_column, dest_column] mappings
        defnname: Name of this definition/report
        busnunitname: Business unit name
        cldrmnth: Calendar month
        cldryear: Calendar year
        totlcols: List of column indices to total
        nzrocols: List of column indices that must be non-zero
        anzrcols: List of column indices for any-non-zero check
        aftrtotldefn: Dictionary of additional totals to add after main totals
        
    Returns:
        int: Number of data rows (headcount) in the populated sheet
    """
    headcntr = 0
    rowsstrt = 5 + len(aftrtotldefn) + 2
    # Gather some details of the Main Excel
    exclrowsused = exclmainshet.max_row
    
    # Add the title to A1
    titlcell = destshet.cell(row = 1, column = 1)
    fontsizelrge(titlcell)
    makefontbold(titlcell)
    titlcell.value = busnunitname + " - " + defnname + " - " + cldrmnth + " " + cldryear
    titlcell.alignment = Alignment(horizontal='general',vertical='center',text_rotation=0,wrap_text=False,shrink_to_fit=False,indent=0)

    destshet.row_dimensions[1].height = None
    
    destcolm = 1
    dlterows = []
    anzrsums = defaultdict(float)
    # Create the Column headers
    for maincolm,thiscolm in thisdefn:

        # Insert the header
        destcell = destshet.cell(row = rowsstrt, column = destcolm)
        destcell.value = thiscolm
        frmttotltitl(destcell)

        # If it is a total column, add the header above also
        if destcolm in totlcols:
            destcell = destshet.cell(row = 1, column = destcolm)
            destcell.value = thiscolm
            frmttotltitl(destcell)

        # Copy the cells below the column headers
        rowscntr = 2
        headcntr = 0
        anzrrows = []
        while rowscntr <= exclrowsused:
            destcell = destshet.cell(row = rowscntr + rowsstrt - 1, column = destcolm)
            destshet.row_dimensions[rowscntr].height = None
            fromcell=exclmainshet.cell(row = rowscntr, column = maincolmhdrs[maincolm])
            destcell.value = fromcell.value 
            fontsizenrml(destcell)

            # Initialize row sum if needed
            if destcell.row not in anzrsums:
                anzrsums[destcell.row] = 0
                
            if destcolm in nzrocols or destcolm in anzrcols:
                if isinstance(destcell.value, (int, float)) and not isinstance(destcell.value, bool):
                    if destcolm in anzrcols:
                        anzrsums[destcell.row] += abs(destcell.value)
                    if destcolm in nzrocols:
                        if abs(destcell.value) < 1e-12:
                            if destcell.row not in dlterows:
                                dlterows.append(destcell.row)
                elif isinstance(destcell.value, str):
                    if destcolm in nzrocols:
                        if destcell.value.strip() == "":
                            if destcell.row not in dlterows:
                                dlterows.append(destcell.row)
                elif destcell.value is None:
                    if destcolm in nzrocols:
                        if destcell.row not in dlterows:
                            dlterows.append(destcell.row)

            # Clean up the values in columns being added up            
            if destcolm in totlcols:
                if destcell.value is None:
                    destcell.value = 0.00
                destcell.number_format = NUMBER_FORMAT

            rowscntr += 1
            headcntr += 1
        destcolm += 1

    if len(totlcols) > 1:
        destcell = destshet.cell(row = rowsstrt, column = destcolm)
        destcell.value = "Total"
        frmttotltitl(destcell)

    # Remove rows that do not have anything in one of the _ANZ_ columns
    # by adding the rows to the rows that should be deleted
    if anzrcols:
        for anzrrown in anzrsums.keys():
            if anzrsums[anzrrown] == 0:
                if anzrrown not in dlterows:
                    dlterows.append(anzrrown)

    # Remove the zero value ones _NZ_ columns from the sheet
    dlterows.sort(reverse = True)
    if dlterows:
        for dlterown in dlterows:
            destshet.delete_rows(dlterown,1)
            headcntr -= 1

    # If all the lines were eliminated, the sheet should not be created
    if headcntr == 0: 
        return headcntr


    # Add up the totals if there are any
    
    if len(totlcols) > 0:
        rowsused = destshet.max_row
        colsused = destshet.max_column
        colmnmbr = 1

        totllinecels = defaultdict(list)

        while colmnmbr <= colsused:

            totlcellblow = destshet.cell(row = rowsused + 1, column = colmnmbr)

            # Colour all the cells in the total line
            fillcellcolr(totlcellblow)

            if colmnmbr in totlcols:
                frmttotlvalu(totlcellblow)

                totlcellabov = destshet.cell(row = 2           , column = colmnmbr)
                frmttotlvalu(totlcellabov)

                ctrlcellabov = destshet.cell(row = 3           , column = colmnmbr)
                #fontsizelrge(ctrlcellabov)
                fontsizenrml(ctrlcellabov)

                rowscntr = rowsstrt + 1
                totlcolmcels = defaultdict(list)
                # ??????????????????????????????????????
                while rowscntr <= rowsused:
                    valucell = destshet.cell(row = rowscntr, column = colmnmbr)
                    if type(valucell.value) in [int, float]:
                        totlcolmcels[colmnmbr].append(valucell.coordinate)
                        totllinecels[rowscntr].append(valucell.coordinate)
                    
                    rowscntr += 1

                # load the cell value into bottom and above cells
                totlcellblow.value = "=SUM(" + totlcolmcels[colmnmbr][0] + ":" + totlcolmcels[colmnmbr][-1] + ")"
                totlcellabov.value = "=SUM(" + totlcolmcels[colmnmbr][0] + ":" + totlcolmcels[colmnmbr][-1] + ")"
                ctrlcellabov.value = "=IF(" + totlcellblow.coordinate+ "=" + totlcellabov.coordinate + ",TRUE,FALSE)"
                
            if colmnmbr == 1:
                totlcellblow.value = "Grand Total"
                frmttotltitl(totlcellblow)

            colmnmbr += 1
   
    rowscntr = rowsstrt + 1 
    rowslast = destshet.max_row - 1
    colmlast = destshet.max_column 
    sidetotllist = []
    
    if len(totlcols) > 1:
        # Row totals on the right
        while rowscntr <= rowslast:
            sidetotlcels = []
            for colmnmbr in totlcols:
                thiscell = destshet.cell(row = rowscntr, column = colmnmbr)
                sidetotlcels.append(thiscell.coordinate)

            thiscell = destshet.cell(row = rowscntr, column = colmlast)
            fontsizenrml(thiscell)
            makefontbold(thiscell)
            thiscell.value = "=SUM(" + ",".join(sidetotlcels) + ")"
            thiscell.number_format = NUMBER_FORMAT
            sidetotllist.append(thiscell.coordinate)
            rowscntr += 1
        
        sidetotlcolm = colsused 
        sidetotlcell = destshet.cell(row = 1, column = sidetotlcolm)
        sidetotlcell.value = "Total"
        frmttotltitl(sidetotlcell)

        # Total on the right Above
        sidetotlabov = destshet.cell(row = 2, column = sidetotlcolm)
        sidetotlabov.value = "=sum(" + sidetotllist[0] + ":" + sidetotllist[-1] + ")"
        frmttotlvalu(sidetotlabov)

        # Bottom total on the right
        sidetotlblow = destshet.cell(row = rowscntr, column = colmlast)
        sidetotlblow.value = "=sum(" + sidetotllist[0] + ":" + sidetotllist[-1] + ")"
        frmttotlvalu(sidetotlblow)

        # Control at top right
        sidectrlcell = destshet.cell( row = 3, column=sidetotlcolm)
        sidectrlcell.value = "=IF(" + sidetotlblow.coordinate+ "=" + sidetotlabov.coordinate + ",TRUE,FALSE)"
        fontsizenrml(sidectrlcell)

    if len(totlcols) > 0:
        # Insert the pesky "Total" before the first total column
        thiscell = destshet.cell( row = 2, column = totlcols[0] - 1 )
        thiscell.value = "Total"
        #fontsizelrge(thiscell)
        fontsizenrml(thiscell)
        makefontbold(thiscell)
        frsttotlcolm=totlcols[0]
    else:
        frsttotlcolm = 6

    # SARS Stuff
    if len(aftrtotldefn) > 0:
        aftrtotlrown = 5
        aftrtotlcels = []
        for aftrtotl in aftrtotldefn:
            aftrtitlcell = destshet.cell( row = aftrtotlrown, column = sidetotlcolm -1 )
            aftrtitlcell.value = aftrtotl.strip("_")
            fontsizenrml(aftrtitlcell)
            makefontbold(aftrtitlcell)
            fillcellcolr(aftrtitlcell)
            
            aftrvalucell = destshet.cell( row = aftrtotlrown, column = sidetotlcolm  )
            # For the sum of sums
            aftrtotlcels.append(aftrvalucell.coordinate)

            # Make a list of coordinates
            aftrsnglcels = []
            for aftrtotlcolm in aftrtotldefn[aftrtotl]:
                aftrsnglcell = destshet.cell( row = 2, column = aftrtotlcolm)
                aftrsnglcels.append(aftrsnglcell.coordinate)
            # Add them up
            aftrvalucell.value = "=SUM(" + ",".join(aftrsnglcels) + ")"
            fontsizenrml(aftrvalucell)
            makefontbold(aftrvalucell)
            aftrvalucell.number_format = NUMBER_FORMAT

            aftrtotlrown = aftrtotlrown + 1
            

        aftrtotltitl = destshet.cell( row = aftrtotlrown, column=sidetotlcolm -1 )
        fontsizenrml(aftrtotltitl)
        makefontbold(aftrtotltitl)
        fillcellcolr(aftrtotltitl)
        aftrtotltitl.value = "Total"
        aftrtotlvalu = destshet.cell( row = aftrtotlrown, column=sidetotlcolm )
        fontsizenrml(aftrtotlvalu)
        makefontbold(aftrtotlvalu)
        aftrtotlvalu.value = "=SUM(" + ",".join(aftrtotlcels) + ")"
        aftrtotlvalu.number_format = NUMBER_FORMAT
        frmttotlvalu(aftrtotlvalu)

            
    # Merge the cells for the title on the left.
    destshet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=frsttotlcolm - 1)

    # Add the headcount on line 2 Column A
    headcntrcell = destshet.cell( row = 2, column = 1 ) 
    headcntrcell.value = 'Total Headcount: ' + str(headcntr)
    makefontbold(headcntrcell)

    # Make all the columns auto sizing
    for cols in destshet.columns:
        col = get_column_letter(cols[0].column)
        destshet.column_dimensions[col].auto_size = True

    return headcntr
    

def processFiles(defnfilename, exclfilename, cldrmnth, cldryear, busnunitname, debug_enabled):
    """Process Excel payroll files according to definition file specifications.
    
    Reads a definition INI file and an Excel payroll file, then creates formatted
    output sheets based on the definitions. Each section in the INI file generates
    a separate worksheet with filtered and formatted data.
    
    Args:
        defnfilename: Path to the INI definition file
        exclfilename: Path to the Excel payroll file
        cldrmnth: Calendar month (e.g., "Jan", "Feb")
        cldryear: Calendar year (e.g., "2025")
        busnunitname: Business unit name
        debug_enabled: Whether to enable debug logging (currently unused, uses logging level)
        
    Returns:
        tuple: (status, result) where status is "Success" or "Failed" and result is 
               empty string on success or error message on failure
    """
    datafilefldr = path.dirname(exclfilename)
    newxfilename = path.join(datafilefldr, str(path.basename(exclfilename.replace(".xlsx", " Tabs.xlsx"))))

    logger.info("Running in folder: %s", os.getcwd())
    logger.info("Selected Month: %s", cldrmnth)
    logger.info("Selected Year: %s", cldryear)
    logger.info("Data Folder: %s", datafilefldr)
    logger.info("Definition file (ini): %s", defnfilename)
    logger.info("Excel Source file: %s", os.path.basename(exclfilename))
    logger.info("New Main Excel file: %s", os.path.basename(newxfilename))
    logger.info("New files created here: %s", datafilefldr)

    # Main, start of the program
    try:

        #Assume all files are present
        filepass = True
   
        # Check if definition file exists
        filechckbool, filechckrslt = check_file(defnfilename)
        if not filechckbool:
            filepass = False

        # Check if Excel file exists
        filechckbool, filechckrslt = check_file(exclfilename)
        if not filechckbool:
            filepass = False

        if not filepass:
            return("Failed","Invalid files selected")

        # Parse the definition file
        defndict = defnfileprse(defnfilename)

        # Open the input Excel sheet and request the result of instead of
        # the formulas itself and take some measurements.
        exclmainbook = load_workbook(exclfilename,data_only=True)
        exclmainshet = exclmainbook[exclmainbook.sheetnames[0]]
        exclcolsused = exclmainshet.max_column
      
        # Make a list of column headers from the input sheet
        colmcntr = 1
        maincolmhdrs = {}
        while colmcntr <= exclcolsused:
            maincolmname=exclmainshet.cell(row = 1, column = colmcntr).value
            maincolmhdrs[maincolmname] = colmcntr
            colmcntr += 1

        # Process each section in the INI file
        for defn in defndict:
            
            givndefn = defndict[defn]
            defnname = defn
            if defnname == "COMPANIES":
                continue
            
            # Create a list of input and output column mappings where the
            # input column exists.
            # nzro - no zeroes allowed
            # anzr - any column msrked with this and has a value will result in
            # inclusion
            # totl These columns will be added up
            thisdefn = list()
            nzrocols = [] 
            anzrcols = []
            totlcols = []
            destcols = defaultdict()

            colmcntr = 1
            aftrtotldefn = defaultdict(list)

            shetslct = True
            anzrpres = False
            for maincolm, thiscolm in givndefn:
                maincolm = maincolm.strip()
                thiscolm = thiscolm.strip()

                # Summed totals to be added after 
                
                if maincolm.startswith("_") and maincolm.endswith("_"):
                    for aftrcolm in thiscolm.split("+"):
                        aftrcolm=aftrcolm.strip()
                        if aftrcolm in maincolmhdrs:
                            if destcols[aftrcolm] in totlcols:
                                aftrtotldefn[maincolm].append(destcols[aftrcolm])
                    

                
                if "_ANZ_" in thiscolm:
                    anzrpres = True
                # Ignore columns that cannot be found in the input sheet.
                if maincolm not in maincolmhdrs:
                    if "_NZ_" in thiscolm:
                        shetslct = False
                        break
                    continue

                # _ANZ_ will have rows with an empty or zero value in this column to to be excluded
                # _ANZ_ is not compatible with _NZ_ and takes priority
                if "_ANZ_" in thiscolm:
                    anzrcols.append(colmcntr)
                    thiscolm = thiscolm.replace("_NZ_", "").strip()
                    thiscolm = thiscolm.replace("_ANZ_", "").strip()
                    
                # _NZ_ will have rows with an empty or zero value in this column to to be excluded
                if "_NZ_" in thiscolm:
                    nzrocols.append(colmcntr)
                    thiscolm = thiscolm.replace("_NZ_", "").strip()

                # _SUM_ will have the values in this column added up and shown below
                if "_SUM_" in thiscolm:
                    totlcols.append(colmcntr)
                    thiscolm = thiscolm.replace("_SUM_", "").strip()
                destcols[maincolm] = colmcntr
                thisdefn.append([maincolm, thiscolm])
                colmcntr += 1

            if anzrpres and len(anzrcols) == 0:
                logger.warning("No ANZ columns found in main sheet for %s", defnname)
                shetslct = False

            logger.debug("Sheet selected: %s for %s", shetslct, defnname)
            if not shetslct:
                continue

            # Create a tab in the copy of the main Excel file
            destshet = exclmainbook.create_sheet(title=defnname)
            destshet.sheet_view.showGridLines = True
            headcntr = populateTheSheet(
                exclmainshet, maincolmhdrs, destshet, thisdefn, defnname,
                busnunitname, cldrmnth, cldryear, totlcols, nzrocols, 
                anzrcols, aftrtotldefn
            )
            logger.info("Sheet '%s' created with headcount: %d", defnname, headcntr)
            
            # Remove the sheet if the report has a zero headcount
            if headcntr <= 0:
                exclmainbook.remove(destshet)
                logger.info("Sheet '%s' removed due to zero headcount", defnname)

    

        # Save the copy with schedules only at the end.
        exclmainbook.save(newxfilename)


    except Exception as e:
        status = "Failed"
        result = e
    else:
        status = "Success"
        result = ""

    return(status, result)
