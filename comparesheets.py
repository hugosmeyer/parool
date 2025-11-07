#!/usr/bin/env python3
"""Excel workbook comparison utility.

This script compares two Excel workbooks for differences in content,
formatting, and structure.
"""

import sys
import openpyxl


def load_workbook(path):
    """Load an Excel workbook with formulas evaluated.
    
    Args:
        path: Path to the Excel file
        
    Returns:
        openpyxl.Workbook: The loaded workbook
    """
    return openpyxl.load_workbook(path, data_only=True)


def compare_cells(cell1, cell2):
    """Compare two cells for equality in value and formatting.
    
    Args:
        cell1: First cell to compare
        cell2: Second cell to compare
        
    Returns:
        bool: True if cells are equal, False otherwise
    """
    # Compare value (literal or evaluated formula result)
    if cell1.value != cell2.value:
        return False

    # Compare font properties
    f1, f2 = cell1.font, cell2.font
    if (
        f1.name != f2.name or
        f1.size != f2.size or
        f1.bold != f2.bold or
        f1.italic != f2.italic
    ):
        return False

    # Compare number format (e.g., date vs plain number)
    if cell1.number_format != cell2.number_format:
        return False

    return True


def sheets_are_equal(sheet1, sheet2):
    """Compare two worksheets for equality.
    
    Args:
        sheet1: First worksheet to compare
        sheet2: Second worksheet to compare
        
    Returns:
        bool: True if sheets are equal, False otherwise
    """
    if sheet1.max_row != sheet2.max_row or sheet1.max_column != sheet2.max_column:
        print(f"Sheet dimensions differ: {sheet1.max_row}x{sheet1.max_column} vs {sheet2.max_row}x{sheet2.max_column}")
        return False

    for row in range(1, sheet1.max_row + 1):
        for col in range(1, sheet1.max_column + 1):
            c1 = sheet1.cell(row, col)
            c2 = sheet2.cell(row, col)
            if not compare_cells(c1, c2):
                print(f"Difference at {c1.coordinate}: '{c1.value}' != '{c2.value}'")
                return False
    return True


def compare_workbooks(path1, path2):
    """Compare two Excel workbooks for equality.
    
    Args:
        path1: Path to first Excel file
        path2: Path to second Excel file
        
    Returns:
        bool: True if workbooks are equal, False otherwise
    """
    try:
        wb1 = load_workbook(path1)
        wb2 = load_workbook(path2)
    except FileNotFoundError as e:
        print(f"Error: {e}")
        return False

    sheets1 = set(wb1.sheetnames)
    sheets2 = set(wb2.sheetnames)

    if sheets1 != sheets2:
        print("Sheet names differ:")
        print("  Only in file1:", sheets1 - sheets2)
        print("  Only in file2:", sheets2 - sheets1)
        return False

    for name in sheets1:
        print(f"Comparing sheet: {name}")
        s1 = wb1[name]
        s2 = wb2[name]
        if not sheets_are_equal(s1, s2):
            print(f"Sheet '{name}' differs.")
            return False

    print("Files are identical in content and formatting.")
    return True


def main():
    """Main entry point for the comparison script."""
    if len(sys.argv) != 3:
        print("Usage: python comparesheets.py file1.xlsx file2.xlsx")
        sys.exit(1)
    
    result = compare_workbooks(sys.argv[1], sys.argv[2])
    sys.exit(0 if result else 1)


if __name__ == "__main__":
    main()

